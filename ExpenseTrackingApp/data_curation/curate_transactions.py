"""Create an app-ready transaction dataset from the legacy expense workbook."""

from __future__ import annotations

import argparse
import calendar
import csv
import hashlib
import json
import re
from collections import Counter, defaultdict
from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any, Iterable

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.worksheet.table import Table, TableStyleInfo


REPO_ROOT = Path(__file__).resolve().parent.parent
DEFAULT_SOURCE = REPO_ROOT / "Spending_Income_Tracker.xlsx"
DEFAULT_OUTPUT_DIR = Path(__file__).resolve().parent / "output"
OUTPUT_WORKBOOK_NAME = "curated_transactions.xlsx"
OUTPUT_CSV_NAME = "curated_transactions.csv"
OUTPUT_REPORT_NAME = "curation_report.json"
ALLOWED_TYPES = ("Income", "Expense", "Investment")
OUTPUT_HEADERS = ("Date", "Type", "Amount", "Description")
CSV_HEADERS = ("occurred_on", "transaction_type", "amount", "description")


@dataclass(frozen=True)
class Transaction:
    transaction_date: date
    transaction_type: str
    amount: Decimal
    description: str
    source_row: int
    source_column: int
    source_month: str


@dataclass(frozen=True)
class MonthBlock:
    start_row: int
    end_row: int
    section_date: date
    expense_header_row: int

    @property
    def label(self) -> str:
        return self.section_date.strftime("%Y-%m")


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def json_value(value: Any) -> Any:
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    if isinstance(value, Decimal):
        return format(value, "f")
    return value


def discover_month_blocks(sheet: openpyxl.worksheet.worksheet.Worksheet) -> list[MonthBlock]:
    headers: list[tuple[int, date]] = []
    for cell in sheet["A"]:
        if isinstance(cell.value, datetime):
            value = cell.value.date()
        elif isinstance(cell.value, date):
            value = cell.value
        else:
            continue
        if cell.number_format.lower() == "mmmm yyyy":
            headers.append((cell.row, value))

    if not headers:
        raise ValueError("No monthly section headers were found.")

    blocks: list[MonthBlock] = []
    for index, (start_row, section_date) in enumerate(headers):
        end_row = headers[index + 1][0] - 1 if index + 1 < len(headers) else sheet.max_row
        expense_headers = [
            row
            for row in range(start_row, end_row + 1)
            if sheet.cell(row, 1).value == "Expenses"
        ]
        if len(expense_headers) != 1:
            raise ValueError(
                f"Expected one Expenses header in {section_date:%Y-%m}; "
                f"found {expense_headers}."
            )
        if sheet.cell(start_row + 1, 1).value != "Income":
            raise ValueError(f"Missing Income header in {section_date:%Y-%m}.")
        if sheet.cell(start_row + 1, 5).value != "Investments":
            raise ValueError(f"Missing Investments header in {section_date:%Y-%m}.")
        blocks.append(
            MonthBlock(
                start_row=start_row,
                end_row=end_row,
                section_date=section_date.replace(day=1),
                expense_header_row=expense_headers[0],
            )
        )
    return blocks


def populated_triplet(sheet: openpyxl.worksheet.worksheet.Worksheet, row: int, start_col: int) -> bool:
    return any(sheet.cell(row, col).value is not None for col in range(start_col, start_col + 3))


def candidate_rows(
    sheet: openpyxl.worksheet.worksheet.Worksheet,
    block: MonthBlock,
) -> list[tuple[str, int, int]]:
    candidates: list[tuple[str, int, int]] = []

    for row in range(block.start_row + 3, block.expense_header_row):
        if populated_triplet(sheet, row, 1):
            candidates.append(("Income", row, 1))

    for row in range(block.start_row + 3, block.end_row + 1):
        if sheet.cell(row, 5).value == "Net Income":
            continue
        if populated_triplet(sheet, row, 5):
            candidates.append(("Investment", row, 5))

    for row in range(block.expense_header_row + 2, block.end_row + 1):
        if populated_triplet(sheet, row, 1):
            candidates.append(("Expense", row, 1))

    return candidates


def normalize_date(raw_value: Any, section_date: date) -> tuple[date, str | None]:
    final_day = calendar.monthrange(section_date.year, section_date.month)[1]
    reason: str | None = None

    if isinstance(raw_value, datetime):
        source_date = raw_value.date()
        day = source_date.day
        if (source_date.year, source_date.month) != (section_date.year, section_date.month):
            reason = "off_section_date"
    elif isinstance(raw_value, date):
        source_date = raw_value
        day = source_date.day
        if (source_date.year, source_date.month) != (section_date.year, section_date.month):
            reason = "off_section_date"
    elif isinstance(raw_value, str):
        day_matches = [
            int(value) for value in re.findall(r"/\s*(\d{1,2})(?!\d)", raw_value)
        ]
        if day_matches:
            day = day_matches[-1]
            reason = "combined_text_date" if len(day_matches) > 1 else "text_date"
        else:
            day = final_day
            reason = "unknown_day_month_end"
    elif raw_value is None:
        day = final_day
        reason = "missing_date_month_end"
    else:
        raise ValueError(f"Unsupported date value {raw_value!r}.")

    if day < 1 or day > final_day:
        day = min(max(day, 1), final_day)
        reason = "invalid_day_clamped"

    normalized = date(section_date.year, section_date.month, day)
    if isinstance(raw_value, (datetime, date)) and normalized == (
        raw_value.date() if isinstance(raw_value, datetime) else raw_value
    ):
        reason = None
    return normalized, reason


def decimal_amount(value: Any, *, cell: str) -> Decimal:
    if isinstance(value, bool) or value is None:
        raise ValueError(f"{cell} does not contain a numeric amount.")
    try:
        amount = Decimal(str(value))
    except (InvalidOperation, ValueError) as exc:
        raise ValueError(f"{cell} contains invalid amount {value!r}.") from exc
    if not amount.is_finite() or amount <= 0:
        raise ValueError(f"{cell} must contain a positive finite amount.")
    if amount.as_tuple().exponent < -2:
        raise ValueError(f"{cell} has more than two decimal places.")
    return amount


def formula_range_rows(formula: Any) -> set[int]:
    if not isinstance(formula, str):
        return set()
    match = re.search(r"[A-Z]+(\d+):[A-Z]+(\d+)", formula, flags=re.IGNORECASE)
    if not match:
        return set()
    return set(range(int(match.group(1)), int(match.group(2)) + 1))


def find_formula_omissions(
    formula_sheet: openpyxl.worksheet.worksheet.Worksheet,
    values_sheet: openpyxl.worksheet.worksheet.Worksheet,
    block: MonthBlock,
    candidates: Iterable[tuple[str, int, int]],
) -> list[dict[str, Any]]:
    formula_cells = {
        "Income": formula_sheet.cell(block.start_row + 1, 3),
        "Investment": formula_sheet.cell(block.start_row + 1, 7),
        "Expense": formula_sheet.cell(block.expense_header_row, 3),
    }
    omissions: list[dict[str, Any]] = []
    for transaction_type, row, start_col in candidates:
        amount = values_sheet.cell(row, start_col).value
        if amount is None:
            continue
        total_cell = formula_cells[transaction_type]
        if row not in formula_range_rows(total_cell.value):
            omissions.append(
                {
                    "source_month": block.label,
                    "type": transaction_type,
                    "total_cell": total_cell.coordinate,
                    "formula": total_cell.value,
                    "omitted_source_row": row,
                    "amount": format(decimal_amount(amount, cell=f"{values_sheet.title}!{values_sheet.cell(row, start_col).coordinate}"), ".2f"),
                    "description": formula_sheet.cell(row, start_col + 2).value,
                }
            )
    return omissions


def extract_transactions(
    formula_sheet: openpyxl.worksheet.worksheet.Worksheet,
    values_sheet: openpyxl.worksheet.worksheet.Worksheet,
) -> tuple[list[Transaction], dict[str, Any]]:
    transactions: list[Transaction] = []
    exclusions: list[dict[str, Any]] = []
    date_normalizations: list[dict[str, Any]] = []
    formula_omissions: list[dict[str, Any]] = []
    candidate_count = 0

    blocks = discover_month_blocks(formula_sheet)
    for block in blocks:
        candidates = candidate_rows(formula_sheet, block)
        candidate_count += len(candidates)
        formula_omissions.extend(
            find_formula_omissions(formula_sheet, values_sheet, block, candidates)
        )

        for transaction_type, row, start_col in candidates:
            raw_amount = values_sheet.cell(row, start_col).value
            raw_date = formula_sheet.cell(row, start_col + 1).value
            raw_description = formula_sheet.cell(row, start_col + 2).value

            if raw_amount is None:
                exclusions.append(
                    {
                        "source_month": block.label,
                        "source_row": row,
                        "reason": "missing_amount",
                        "date": json_value(raw_date),
                        "description": raw_description,
                    }
                )
                continue

            amount = decimal_amount(
                raw_amount,
                cell=f"{values_sheet.title}!{values_sheet.cell(row, start_col).coordinate}",
            )
            if raw_description is None or not str(raw_description).strip():
                raise ValueError(
                    f"{formula_sheet.title}!{formula_sheet.cell(row, start_col + 2).coordinate} "
                    "does not contain a description."
                )
            description = str(raw_description).strip()
            normalized_date, date_reason = normalize_date(raw_date, block.section_date)
            if date_reason:
                date_normalizations.append(
                    {
                        "source_month": block.label,
                        "source_row": row,
                        "source_cell": formula_sheet.cell(row, start_col + 1).coordinate,
                        "reason": date_reason,
                        "original": json_value(raw_date),
                        "normalized": normalized_date.isoformat(),
                    }
                )

            transactions.append(
                Transaction(
                    transaction_date=normalized_date,
                    transaction_type=transaction_type,
                    amount=amount,
                    description=description,
                    source_row=row,
                    source_column=start_col,
                    source_month=block.label,
                )
            )

    transactions.sort(
        key=lambda item: (
            item.transaction_date,
            item.source_row,
            item.source_column,
        )
    )
    duplicate_groups: list[dict[str, Any]] = []
    grouped: dict[tuple[Any, ...], list[Transaction]] = defaultdict(list)
    for transaction in transactions:
        grouped[
            (
                transaction.transaction_date,
                transaction.transaction_type,
                transaction.amount,
                transaction.description,
            )
        ].append(transaction)
    for key, group in grouped.items():
        if len(group) > 1:
            duplicate_groups.append(
                {
                    "date": key[0].isoformat(),
                    "type": key[1],
                    "amount": format(key[2], ".2f"),
                    "description": key[3],
                    "source_rows": [item.source_row for item in group],
                }
            )

    details = {
        "month_blocks": len(blocks),
        "candidate_records": candidate_count,
        "exclusions": exclusions,
        "date_normalizations": date_normalizations,
        "formula_omissions": formula_omissions,
        "duplicate_groups": duplicate_groups,
    }
    return transactions, details


def write_workbook(transactions: list[Transaction], output_path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Transactions"
    sheet.sheet_view.showGridLines = False
    sheet.freeze_panes = "A2"
    sheet.append(OUTPUT_HEADERS)

    for transaction in transactions:
        sheet.append(
            (
                transaction.transaction_date,
                transaction.transaction_type,
                float(transaction.amount),
                transaction.description,
            )
        )

    table_ref = f"A1:D{len(transactions) + 1}"
    table = Table(displayName="TransactionsTable", ref=table_ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    sheet.add_table(table)

    for cell in sheet[1]:
        cell.font = Font(name="Aptos", size=11, bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="left", vertical="center")
    sheet.row_dimensions[1].height = 22

    for row in range(2, len(transactions) + 2):
        sheet.cell(row, 1).number_format = "yyyy-mm-dd"
        sheet.cell(row, 1).alignment = Alignment(horizontal="left", vertical="center")
        sheet.cell(row, 2).alignment = Alignment(horizontal="left", vertical="center")
        sheet.cell(row, 3).number_format = "#,##0.00"
        sheet.cell(row, 3).alignment = Alignment(horizontal="right", vertical="center")
        sheet.cell(row, 4).alignment = Alignment(horizontal="left", vertical="center")
        for column in range(1, 5):
            sheet.cell(row, column).font = Font(name="Aptos", size=11)
        sheet.row_dimensions[row].height = 18

    widths = {"A": 13, "B": 14, "C": 14, "D": 32}
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width

    sheet.auto_filter.ref = table_ref
    workbook.properties.creator = "ExpenseTrackingApp data curation"
    workbook.properties.title = "Curated spending and income transactions"
    workbook.properties.subject = "App-ready transaction dataset"
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)
    workbook.close()


def write_csv(transactions: list[Transaction], output_path: Path) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with output_path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.writer(handle, lineterminator="\n")
        writer.writerow(CSV_HEADERS)
        for transaction in transactions:
            writer.writerow(
                (
                    transaction.transaction_date.isoformat(),
                    transaction.transaction_type.lower(),
                    format(transaction.amount, ".2f"),
                    transaction.description,
                )
            )


def counts_by_month(transactions: list[Transaction]) -> dict[str, dict[str, int]]:
    result: dict[str, Counter[str]] = defaultdict(Counter)
    for transaction in transactions:
        result[transaction.source_month][transaction.transaction_type] += 1
    return {
        month: {kind: result[month].get(kind, 0) for kind in ALLOWED_TYPES}
        for month in sorted(result)
    }


def validate_transactions(transactions: list[Transaction]) -> dict[str, Any]:
    errors: list[str] = []
    for index, transaction in enumerate(transactions, start=2):
        if transaction.transaction_type not in ALLOWED_TYPES:
            errors.append(f"Row {index}: invalid type {transaction.transaction_type!r}.")
        if not isinstance(transaction.transaction_date, date):
            errors.append(f"Row {index}: date is not typed.")
        if transaction.amount <= 0 or transaction.amount.as_tuple().exponent < -2:
            errors.append(f"Row {index}: invalid amount {transaction.amount}.")
        if not transaction.description:
            errors.append(f"Row {index}: description is blank.")
    return {
        "passed": not errors,
        "errors": errors,
        "rules": [
            "allowed transaction type",
            "typed date",
            "positive amount with at most two decimals",
            "nonblank description",
        ],
    }


def curate(source_path: Path = DEFAULT_SOURCE, output_dir: Path = DEFAULT_OUTPUT_DIR) -> dict[str, Any]:
    source_path = source_path.resolve()
    output_dir = output_dir.resolve()
    if not source_path.is_file():
        raise FileNotFoundError(f"Source workbook not found: {source_path}")

    source_hash_before = sha256_file(source_path)
    formula_workbook = openpyxl.load_workbook(source_path, data_only=False)
    values_workbook = openpyxl.load_workbook(source_path, data_only=True)
    try:
        if formula_workbook.sheetnames != ["Sheet1"]:
            raise ValueError(f"Unexpected worksheet layout: {formula_workbook.sheetnames}")
        transactions, details = extract_transactions(
            formula_workbook["Sheet1"],
            values_workbook["Sheet1"],
        )
    finally:
        formula_workbook.close()
        values_workbook.close()

    validation = validate_transactions(transactions)
    if not validation["passed"]:
        raise ValueError(f"Curated transaction validation failed: {validation['errors']}")

    workbook_path = output_dir / OUTPUT_WORKBOOK_NAME
    csv_path = output_dir / OUTPUT_CSV_NAME
    report_path = output_dir / OUTPUT_REPORT_NAME
    write_workbook(transactions, workbook_path)
    write_csv(transactions, csv_path)

    counts = Counter(item.transaction_type for item in transactions)
    totals = {
        kind: format(
            sum(
                (item.amount for item in transactions if item.transaction_type == kind),
                Decimal("0"),
            ),
            ".2f",
        )
        for kind in ALLOWED_TYPES
    }
    source_hash_after = sha256_file(source_path)
    report = {
        "version": 2,
        "source_file": source_path.name,
        "output_file": workbook_path.name,
        "csv_output_file": csv_path.name,
        "csv_sha256": sha256_file(csv_path),
        "source_sha256_before": source_hash_before,
        "source_sha256_after": source_hash_after,
        "source_unchanged": source_hash_before == source_hash_after,
        "month_blocks": details["month_blocks"],
        "candidate_records": details["candidate_records"],
        "curated_records": len(transactions),
        "excluded_records": len(details["exclusions"]),
        "counts_by_type": {kind: counts.get(kind, 0) for kind in ALLOWED_TYPES},
        "counts_by_month": counts_by_month(transactions),
        "totals_by_type": totals,
        "date_normalization_count": len(details["date_normalizations"]),
        "date_normalizations": details["date_normalizations"],
        "exclusions": details["exclusions"],
        "formula_omissions": details["formula_omissions"],
        "duplicate_groups": details["duplicate_groups"],
        "validation": validation,
    }
    if not report["source_unchanged"]:
        raise RuntimeError("Source workbook changed during curation.")

    output_dir.mkdir(parents=True, exist_ok=True)
    report_path.write_text(
        json.dumps(report, indent=2, sort_keys=True, ensure_ascii=False) + "\n",
        encoding="utf-8",
    )
    return report


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Curate the legacy expense workbook into one transaction table."
    )
    parser.add_argument(
        "--source",
        type=Path,
        default=DEFAULT_SOURCE,
        help=f"Source workbook (default: {DEFAULT_SOURCE})",
    )
    parser.add_argument(
        "--output-dir",
        type=Path,
        default=DEFAULT_OUTPUT_DIR,
        help=f"Generated output directory (default: {DEFAULT_OUTPUT_DIR})",
    )
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    report = curate(args.source, args.output_dir)
    print(
        f"Created {report['output_file']} and {report['csv_output_file']} with "
        f"{report['curated_records']} records; report written to {OUTPUT_REPORT_NAME}."
    )


if __name__ == "__main__":
    main()
