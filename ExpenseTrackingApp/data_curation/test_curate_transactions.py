"""Tests for the legacy workbook curation workflow."""

from __future__ import annotations

import csv
import json
import tempfile
import unittest
import zipfile
from datetime import date
from decimal import Decimal
from pathlib import Path

import openpyxl

from data_curation.curate_transactions import (
    CSV_HEADERS,
    DEFAULT_SOURCE,
    OUTPUT_CSV_NAME,
    OUTPUT_REPORT_NAME,
    OUTPUT_WORKBOOK_NAME,
    Transaction,
    curate,
    normalize_date,
    sha256_file,
    write_csv,
)


EXPECTED_SOURCE_SHA256 = "3bf0d327144484b53c3717d26689e585a30d631a1faebe0bd125a51cd531e7a0"


class CurationTests(unittest.TestCase):
    @classmethod
    def setUpClass(cls) -> None:
        cls.temp_dir = tempfile.TemporaryDirectory()
        cls.output_dir = Path(cls.temp_dir.name)
        cls.source_hash_before = sha256_file(DEFAULT_SOURCE)
        cls.report = curate(DEFAULT_SOURCE, cls.output_dir)
        cls.source_hash_after = sha256_file(DEFAULT_SOURCE)
        cls.output_workbook = cls.output_dir / OUTPUT_WORKBOOK_NAME
        cls.output_csv = cls.output_dir / OUTPUT_CSV_NAME
        cls.output_report = cls.output_dir / OUTPUT_REPORT_NAME

    @classmethod
    def tearDownClass(cls) -> None:
        cls.temp_dir.cleanup()

    def test_source_is_unchanged(self) -> None:
        self.assertEqual(self.source_hash_before, EXPECTED_SOURCE_SHA256)
        self.assertEqual(self.source_hash_before, self.source_hash_after)
        self.assertTrue(self.report["source_unchanged"])

    def test_counts_and_totals(self) -> None:
        self.assertEqual(self.report["month_blocks"], 25)
        self.assertEqual(self.report["candidate_records"], 456)
        self.assertEqual(self.report["curated_records"], 452)
        self.assertEqual(self.report["excluded_records"], 4)
        self.assertEqual(
            self.report["counts_by_type"],
            {"Income": 30, "Expense": 340, "Investment": 82},
        )
        self.assertEqual(
            self.report["totals_by_type"],
            {
                "Income": "78632.30",
                "Expense": "10696.35",
                "Investment": "66824.56",
            },
        )

    def test_known_formula_omissions_are_included(self) -> None:
        omissions = {
            (item["source_month"], item["omitted_source_row"], item["amount"])
            for item in self.report["formula_omissions"]
        }
        self.assertEqual(
            omissions,
            {
                ("2025-10", 534, "248.21"),
                ("2025-11", 566, "345.50"),
            },
        )

    def test_date_normalization_rules(self) -> None:
        self.assertEqual(normalize_date("9/??", date(2024, 9, 1))[0], date(2024, 9, 30))
        self.assertEqual(
            normalize_date("5/25, 5/31", date(2026, 5, 1))[0],
            date(2026, 5, 31),
        )
        self.assertEqual(
            normalize_date("3/13,3/27", date(2026, 3, 1))[0],
            date(2026, 3, 27),
        )
        self.assertEqual(normalize_date("6/15", date(2026, 7, 1))[0], date(2026, 7, 15))
        self.assertEqual(
            normalize_date(date(2026, 12, 24), date(2025, 12, 1))[0],
            date(2025, 12, 24),
        )
        self.assertEqual(self.report["date_normalization_count"], 102)

    def test_exclusions_and_duplicates(self) -> None:
        self.assertEqual(
            [item["source_row"] for item in self.report["exclusions"]],
            [37, 39, 40, 41],
        )
        self.assertEqual(len(self.report["duplicate_groups"]), 3)

    def test_workbook_structure_and_values(self) -> None:
        workbook = openpyxl.load_workbook(self.output_workbook, data_only=False)
        try:
            self.assertEqual(workbook.sheetnames, ["Transactions"])
            sheet = workbook["Transactions"]
            self.assertEqual(sheet.max_row, 453)
            self.assertEqual(sheet.max_column, 4)
            self.assertEqual(
                tuple(sheet.cell(1, column).value for column in range(1, 5)),
                ("Date", "Type", "Amount", "Description"),
            )
            self.assertEqual(sheet.freeze_panes, "A2")
            self.assertEqual(set(sheet.tables), {"TransactionsTable"})
            self.assertEqual(sheet.tables["TransactionsTable"].ref, "A1:D453")
            self.assertFalse(sheet.merged_cells.ranges)
            self.assertEqual(
                [
                    cell.coordinate
                    for row in sheet.iter_rows()
                    for cell in row
                    if cell.data_type == "f"
                ],
                [],
            )

            counts: dict[str, int] = {"Income": 0, "Expense": 0, "Investment": 0}
            totals: dict[str, Decimal] = {
                "Income": Decimal("0"),
                "Expense": Decimal("0"),
                "Investment": Decimal("0"),
            }
            for row in sheet.iter_rows(min_row=2, values_only=False):
                transaction_date, transaction_type, amount, description = row
                self.assertIsInstance(transaction_date.value, (date,))
                self.assertIn(transaction_type.value, counts)
                self.assertIsInstance(amount.value, (int, float))
                self.assertGreater(amount.value, 0)
                self.assertTrue(str(description.value).strip())
                self.assertEqual(transaction_date.number_format, "yyyy-mm-dd")
                self.assertEqual(amount.number_format, "#,##0.00")
                counts[transaction_type.value] += 1
                totals[transaction_type.value] += Decimal(str(amount.value))

            self.assertEqual(
                counts,
                {"Income": 30, "Expense": 340, "Investment": 82},
            )
            self.assertEqual(
                {key: format(value, ".2f") for key, value in totals.items()},
                {
                    "Income": "78632.30",
                    "Expense": "10696.35",
                    "Investment": "66824.56",
                },
            )
        finally:
            workbook.close()

    def test_formula_amounts_are_materialized(self) -> None:
        workbook = openpyxl.load_workbook(self.output_workbook, data_only=False)
        try:
            sheet = workbook["Transactions"]
            rows = {
                (
                    row[0].value.date(),
                    row[1].value,
                    Decimal(str(row[2].value)),
                    row[3].value,
                )
                for row in sheet.iter_rows(min_row=2, max_col=4)
            }
            self.assertIn((date(2025, 7, 3), "Income", Decimal("2328.51"), "Paycheck"), rows)
            self.assertIn((date(2026, 4, 24), "Expense", Decimal("1.96"), "Dental"), rows)
            self.assertIn((date(2026, 4, 24), "Expense", Decimal("0.3"), "Vision"), rows)
            self.assertIn(
                (date(2026, 4, 24), "Expense", Decimal("36.44"), "Health Insurance"),
                rows,
            )
        finally:
            workbook.close()

    def test_csv_structure_and_values(self) -> None:
        with self.output_csv.open(encoding="utf-8", newline="") as handle:
            reader = csv.DictReader(handle)
            self.assertEqual(tuple(reader.fieldnames or ()), CSV_HEADERS)
            rows = list(reader)

        self.assertEqual(len(rows), 452)
        counts = {"income": 0, "expense": 0, "investment": 0}
        totals = {
            "income": Decimal("0"),
            "expense": Decimal("0"),
            "investment": Decimal("0"),
        }
        for row in rows:
            date.fromisoformat(row["occurred_on"])
            self.assertIn(row["transaction_type"], counts)
            amount = Decimal(row["amount"])
            self.assertGreater(amount, 0)
            self.assertEqual(amount.as_tuple().exponent, -2)
            self.assertTrue(row["description"].strip())
            counts[row["transaction_type"]] += 1
            totals[row["transaction_type"]] += amount

        self.assertEqual(
            counts,
            {"income": 30, "expense": 340, "investment": 82},
        )
        self.assertEqual(
            {key: format(value, ".2f") for key, value in totals.items()},
            {
                "income": "78632.30",
                "expense": "10696.35",
                "investment": "66824.56",
            },
        )

    def test_csv_matches_workbook_order(self) -> None:
        with self.output_csv.open(encoding="utf-8", newline="") as handle:
            csv_rows = list(csv.DictReader(handle))

        workbook = openpyxl.load_workbook(self.output_workbook, data_only=True)
        try:
            workbook_rows = [
                {
                    "occurred_on": row[0].value.date().isoformat(),
                    "transaction_type": row[1].value.lower(),
                    "amount": format(Decimal(str(row[2].value)), ".2f"),
                    "description": row[3].value,
                }
                for row in workbook["Transactions"].iter_rows(min_row=2, max_col=4)
            ]
        finally:
            workbook.close()

        self.assertEqual(csv_rows, workbook_rows)

    def test_csv_escapes_special_descriptions(self) -> None:
        output_path = self.output_dir / "csv_escaping_test.csv"
        description = 'Transit, "express"\nsecond line'
        write_csv(
            [
                Transaction(
                    transaction_date=date(2026, 7, 26),
                    transaction_type="Expense",
                    amount=Decimal("12.50"),
                    description=description,
                    source_row=1,
                    source_column=1,
                    source_month="2026-07",
                )
            ],
            output_path,
        )

        with output_path.open(encoding="utf-8", newline="") as handle:
            rows = list(csv.DictReader(handle))
        self.assertEqual(rows[0]["description"], description)

    def test_outputs_are_valid_files(self) -> None:
        self.assertTrue(self.output_workbook.is_file())
        self.assertTrue(self.output_csv.is_file())
        self.assertTrue(self.output_report.is_file())
        with zipfile.ZipFile(self.output_workbook) as archive:
            self.assertIsNone(archive.testzip())
        report_from_disk = json.loads(self.output_report.read_text(encoding="utf-8"))
        self.assertEqual(report_from_disk, self.report)
        self.assertEqual(self.report["version"], 2)
        self.assertEqual(self.report["csv_output_file"], OUTPUT_CSV_NAME)
        self.assertEqual(self.report["csv_sha256"], sha256_file(self.output_csv))
        self.assertTrue(self.report["validation"]["passed"])


if __name__ == "__main__":
    unittest.main()
